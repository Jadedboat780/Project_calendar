import googleapiclient
from google.oauth2 import service_account
from googleapiclient.discovery import build
import requests
from bs4 import BeautifulSoup
from datetime import date, datetime
import openpyxl
import schedule
import re
import telebot
from id import name_groups, name_teacher, CalendarID, Token_bot, TG_Bot_ID

# Создаем подключение к календарю указывая для него разрешение доступа
SCOPES = ['https://www.googleapis.com/auth/calendar']
calendarId = '35fc2b53f41517c857c7cd03942345f7ae6770f333eb04e36c768d03ce453e04@group.calendar.google.com'
SERVICE_ACCOUNT_FILE = 'TESTS.json'
credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = googleapiclient.discovery.build('calendar', 'v3', credentials=credentials)

# Создаем переменную для работы с тг ботом
bot = telebot.TeleBot(Token_bot)
id_events = list()  # Храним id ивентов которые надо удалить
Name_list = list()  # Хранит список всех преподователей и групп из планшетки


def start_message(message: str, text: str):
    """Функция отправляет админам сообщения об ошибки"""
    for i in message:
        bot.send_message(i, f"{text}")


def New_Events(text: str, description: str, day: str, time: str, color: str, group: str):  # Метод для загрузки данных в календарь для преподователей
    """Функция загружает данные в календарь"""
    if '-' not in time:
        time = time.split(' — ')
    else:
        time = time.split('-')

    time[0] = time[0].replace(' ', '')
    time[-1] = time[-1].replace(' ', '')
    event = {
        'summary': text,
        'description': description,
        'colorId': color,
        'start': {
            'dateTime': f'{day}T{time[0]}:00+03:00',
        },
        'end': {
            'dateTime': f'{day}T{time[-1]}:00+03:00',
        }}
    try:
        service.events().insert(calendarId=CalendarID[group], body=event).execute()  # Выполняем запрос о создании ивента
        print(f'Был создан ивент - {group}')
    except Exception as e:
        print('Error')
        start_message(TG_Bot_ID, f'Произошла ошибка \n'
                                 f'ошибка была вызвана этими данными: \n \n'
                                 f'text = {text}, \ndescription = {description}, \nday = {day}, \ntime = {time}, \ncolor = {color}, \ngroup = {group} \n'
                                 f'описание ошибки: {e}')


def Delete_Events():
    """Функция удаляет найденные ивенты в календаре"""
    num = -1  # Создаем переменную для поиска id в списке
    for i in id_events:  # Проходимся по списку
        try:
            num += 1  # Сразу же увеличиваем ее
            service.events().delete(calendarId=CalendarID[i], eventId=id_events[num + 1]).execute()  # Находим нужный
        except:
            pass

    id_events.clear()
    Parsing.Planchette()  # Запускаем функцию парсинга планшеток


def get_events_list(time: str):  # Принимает аргумент дата из планшетки
    """Функция находит id ивентов в календаре и записывает их в лист"""
    id_events.clear()  # Очищаем список
    for i in CalendarID:
        n = i.strip()
        if n in name_teacher or n in name_groups:
            try:
                now = datetime.utcnow().isoformat() + 'Z'
                events_result = service.events().list(calendarId=CalendarID[n], timeMin=now, maxResults=10,
                                                      singleEvents=True, orderBy='startTime').execute()
                events = events_result.get('items', [])  # Ищем ивенты
                for event in events:  # Проходимся по найденным ивентам
                    start = event['start'].get('dateTime')  # Записываем в переменную время ивента
                    start = start.split('T')  # Разделяем на дату и время
                    strs = event.get('id')  # Получаем id ивента
                    if start[0] == time:  # если дата в ивенте == дате в планшетке
                        id_events.append(n)  # Записываем в список у какой группы мы парсили календарь
                        id_events.append(strs)  # добавляем в список id ивентов которые надо удалить
            except Exception as e:
                start_message(TG_Bot_ID, f'Была вызвана ошибка в разделе поиска id ивентов \n'
                                         f'Данные вызванные ошибку: {i} \n'
                                         f'описание ошибки: {e}')
        else:
            start_message(TG_Bot_ID, f'Такого преподователя или группы нет | раздел поиск ивентов в календаре \n'
                                     f'Данные которых нет в списке: \'{i}\' -- \'{n}\'')

    Name_list.clear()
    Delete_Events()


def planchet_id():
    """Ишем id планшеток по нынешной дате и запускае функцию ее скачивания"""
    start_message(TG_Bot_ID, f'Парсинг расписания планшетки начал работу')
    try:
        response = requests.get(url='https://drive.google.com/drive/folders/1UH5pcJc0pxkYFWBMI_bNCxrdlApqv3nX')
        soup = BeautifulSoup(response.text, 'lxml')
        soup = soup.find_all('div', class_='Q5txwe')  # Достаем название планшетки

        for number, i in enumerate(soup, start=0):  # Проходимся по всем найденным планшеткам
            dates = str(datetime.now()).split()[0]
            dts = dates.split('-')
            dates = f'{dts[-1]}.{dts[1]}.{dts[0]}'  # Преобразуем текущую дату в нужный вид
            plans = i.text[0:-5]  # Убираем с названия планшетки '.xlsx'

            if dates == plans:  # Если текущая дата совпадает с датой планшетки
                response = requests.get(url='https://drive.google.com/drive/folders/1UH5pcJc0pxkYFWBMI_bNCxrdlApqv3nX')
                soup = BeautifulSoup(response.text, 'lxml')
                id_list = [i.replace('data-id="', '') for i in
                           re.findall(r'data-id="[\w\-\+/#\(\)\*&\^:<>\?\!%\$]+', str(soup))]
                data_exel(identifier=id_list[number])  # Запускаем скачивание планшетки по ее id
                break
    except Exception as e:
        start_message(TG_Bot_ID, f'Была вызвана ошибка в разделе поиска  планшеток \n'
                                 f'описание ошибки: {e} \n'
                                 f'!ИЗ-ЗА ОШИБКИ ДАЛЬНЕЙШАЯ РАБОТА ПРОГРАММЫ ОСТАНОВЛЕНА ДО ПОВТОРНОГО ЕГО ВЫЗОВА!')


def data_exel(identifier: str):  # Функция скачивания планшетки
    """Функция скачивает найденную планшетку"""
    try:
        planchet = requests.get(url=f'https://drive.google.com/uc?export=download&id={identifier}')
        with open('planchette.xlsx', 'wb') as xlsx_file:
            xlsx_file.write(planchet.content)
        Parsing.planchette_check()  # Запускаем парсинг планшетки
    except Exception as e:
        start_message(TG_Bot_ID, 'Была вызвана ошибка в разделе скачивания планшетки \n'
                                 f'описание ошибки: {e} \n'
                                 f'!ИЗ-ЗА ОШИБКИ ДАЛЬНЕЙШАЯ РАБОТА ПРОГРАММЫ ОСТАНОВЛЕНА ДО ПОВТОРНОГО ЕГО ВЫЗОВА!')


class Parsing:  # Запускаем этот класс раз в неделю
    """Класс содержит функции для парсинга и обработки данных"""

    def Par_Techer():  # Парсинг расписания преподователей
        """Функция достает расписание преподователей с сайта, обрабатывает и
        вызывает функцию для загрузки их в календарь"""
        print('Начал свою работу')
        url = 'https://rksi.ru/schedule'
        res = requests.get(url)
        soup = BeautifulSoup(res.text, 'lxml')
        lecturer = soup.find('select', id='teacher').find_all('option')
        clear_list = []
        preps = []
        for i in lecturer:
            clear_list.append(i.text)  # Засовываем все группы в список

        for teacher in name_teacher:
            if teacher in clear_list:
                trash_response_text = requests.post(url, {'teacher': f'{teacher}'.encode('cp1251'),
                                                          'stp': 'Показать!'.encode('cp1251')}).text
                trash_prep_soup = BeautifulSoup(trash_response_text, 'lxml')

                teachers_list = []
                for i in trash_prep_soup.find_all(['p', 'b']):
                    teachers_list.append(
                        str(i).replace('<br/><b>', '||').replace('</b><br/>', '||').replace('<p>', '').replace('<b>', '').replace('</p>', '').replace('</b>', ''))

                preps.append(f'&{teacher}')
                for i in teachers_list:
                    if '</' not in i:
                        preps.append(i.split('||'))

        today = str(date.today())  # Текущая дата
        days = list()  # Засовываем дату (год-месяц-день) текущей пары
        groups = str()  # Для хранения группы
        for name in preps:  # Обрабатываем данные
            if name[0] == "&":
                groups = f"{name[1:]}"
            if 15 <= len(name[0]) <= 23 and name[0][0:1].isdigit():
                nd = name[0].split(" ")  # Разделяем дату на 'день', 'месяц', 'день недели'
                nd[1] = nd[1].replace(",", "")  # Избавляемся от знака , в месяце и засовываем в список
                month = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04', 'мая': '05', 'июня': '06',
                         'июля': '07', 'августа': '08', 'сентября': '09', 'октября': '10', 'ноября': '11',
                         'декабря': '12'}
                if len(nd[0]) == 1 and 1 <= int(nd[0]) < 10:
                    days.append(f"{today[0:4]}-{month[nd[1]]}-0{nd[0]}")
                elif len(nd[0]) == 2 and int(nd[0]) >= 10:
                    days.append(f"{today[0:4]}-{month[nd[1]]}-{nd[0]}")
            if len(name) == 3:  # Обработанные данные загружаем в календарь
                #print(f'{groups} -- {name[1]} - Время: {name[0]}: {name[2]} день {days[-1]}')
                New_Events(f'{name[1]}', f'Время: {name[0]}: {name[2]}', f'{days[-1]}', f'{name[0]}', '9', groups)

    def Par_Group():  # Парсинг расписания групп
        """Функция достает расписание групп с сайта, обрабатывает и вызывает функцию для загрузки их в календарь"""
        url = 'https://rksi.ru/schedule'

        res = requests.get(url)
        soup = BeautifulSoup(res.text, 'lxml')
        lecturer = soup.find('select', id='group').find_all('option')
        clear_list = []
        for i in lecturer:
            clear_list.append(i.text)  # Засовываем все группы в список

        preps = []

        for group in name_groups:
            if group in clear_list:
                trash_response_text = requests.post(url, {'group': f'{group}'.encode('cp1251'),
                                                          'stt': 'Показать!'.encode('cp1251')}).text
                trash_prep_soup = BeautifulSoup(trash_response_text, 'lxml')

                group_list = []
                for i in trash_prep_soup.find_all(['p', 'b']):
                    group_list.append(
                        str(i).replace('<br/><b>', '||').replace('</b><br/>', '||').replace('<p>', '').replace('<b>',
                                                                                                               '').replace(
                            '</p>', '').replace('</b>', ''))

                preps.append(f'&{group}')
                for i in group_list:
                    if '</' not in i:
                        preps.append(i.split('||'))

        today = str(date.today())  # Текущая дата
        days = list()  # Засовываем дату (год-месяц-день) текущей пары
        groups = str()  # Для хранения группы

        for name in preps:  # Обрабатываем данные
            if name[0] == "&":
                groups = f"{name[1:]}"
            if 15 <= len(name[0]) <= 23 and name[0][0:1].isdigit():
                nd = name[0].split(" ")  # Разделяем дату на 'день', 'месяц', 'день недели'
                nd[1] = nd[1].replace(",", "")  # Избавляемся от знака , в месяце и засовываем в список
                month = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04', 'мая': '05', 'июня': '06',
                         'июля': '07', 'августа': '08', 'сентября': '09', 'октября': '10', 'ноября': '11',
                         'декабря': '12'}
                if len(nd[0]) == 1 and 1 <= int(nd[0]) < 10:
                    days.append(f"{today[0:4]}-{month[nd[1]]}-0{nd[0]}")
                elif len(nd[0]) == 2 and int(nd[0]) >= 10:
                    days.append(f"{today[0:4]}-{month[nd[1]]}-{nd[0]}")

            if len(name) == 3:  # Обработанные данные загружаем в календарь
                #print(f'{groups} -- {name[1]} - Время: {name[0]}: {name[2]} день {days[-1]}')
                New_Events(f'{name[1]}', f'Время: {name[0]}: {name[2]}', f'{days[-1]}', f'{name[0]}', '9', groups)

    def planshet():
        """Функция достает и обрабатывает данные с планшетки и запускает функцию удаления ивентов"""
        wb1 = openpyxl.load_workbook('planchette.xlsx')
        couple = '1 пара'

        worksheet = wb1[couple]  # Открываем листы в планшетке
        dates = str(worksheet['B1'].value).split()  # Достаем дату

        if '-' in dates[0]:  # Разделяем дату
            dates = dates[0].split('-')
        else:
            dates = dates[0].split('.')
        dates = f'{dates[2]}-{dates[1]}-{dates[0]}'  # Записываем дату в нужном нам формате
        get_events_list(dates)  # Вызываем функцию поиска ивентов

    def Planchette():
        """Функция достает и обрабатывает данные с планшетки и запускает функцию загрузки их в календарь"""
        wb1 = openpyxl.load_workbook('planchette.xlsx')
        days = datetime.today().isoweekday()
        roster = list()

        if days == 1:
            couple_list = ['1 пара', '2 пара', '3 пара', 'Классный час', '4 пара', '5 пара', '6 пара']
        else:
            couple_list = ['1 пара', '2 пара', '3 пара', '4 пара', '5 пара', '6 пара', '7 пара']

        try:
            for couple in couple_list:
                worksheet = wb1[couple]
                dates = str(worksheet['B1'].value).split()
                if '-' in dates[0]:
                    dates = dates[0].split('-')
                else:
                    dates = dates[0].split('.')
                dates1 = f'{dates[2]}-{dates[1]}-{dates[0]}'
                couple_time = worksheet['A1'].value
                for i in range(0, worksheet.max_row):
                    my_list_1 = []
                    my_list_2 = []
                    for col in worksheet.iter_cols(2, 5):
                        a = col[i].value
                        if a is not None:
                            my_list_1.append(a)

                    if len(my_list_1) == 3:
                        if '/' in my_list_1[1]:
                            for j in my_list_1[1].split('/'):
                                if 'КПК' not in my_list_1 and f'{dates[0]}.{dates[1]}.{dates[2]}' not in my_list_1:
                                    roster.append([[couple_time],
                                                   ['Без названия', f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}',
                                                    dates1, couple_time, 7, my_list_2[1]]])
                        else:
                            if 'КПК' not in my_list_1 and f'{dates[0]}.{dates[1]}.{dates[2]}' not in my_list_1:
                                roster.append([[couple_time],
                                               ['Без названия', f'{couple_time}: {my_list_1[2]} ауд {my_list_1[0]}',
                                                dates1, couple_time, 7, my_list_1[1]]])

                    elif len(my_list_1) >= 4:
                        if '/' in my_list_1[1]:
                            for j in my_list_1[1].split('/'):
                                roster.append(
                                    [couple_time, my_list_1[3], f'{couple_time}: {my_list_1[2]} ауд {my_list_1[0]}',
                                     dates1, couple_time, 9, j])
                        else:
                            roster.append([[couple_time],
                                           [my_list_1[3], f'{couple_time}: {my_list_1[2]} ауд {my_list_1[0]}', dates1,
                                            couple_time, 9, my_list_1[1]]])

                    for col in worksheet.iter_cols(6, 9):
                        a = col[i].value

                        if a is not None:
                            my_list_2.append(a)

                    if len(my_list_2) == 3:
                        if '/' in my_list_2[1]:
                            for j in my_list_1[2].split('/'):
                                if 'КПК' not in my_list_2 and f'{dates[0]}.{dates[1]}.{dates[2]}' not in my_list_2:
                                    roster.append([[couple_time],
                                                   ['Без названия', f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}',
                                                    dates1, couple_time, 7, my_list_2[1]]])

                        else:
                            if 'КПК' not in my_list_2 and f'{dates[0]}.{dates[1]}.{dates[2]}' not in my_list_2:
                                roster.append([[couple_time],
                                               ['Без названия', f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}',
                                                dates1, couple_time, 7, my_list_2[1]]])

                    elif len(my_list_2) >= 4:
                        if '/' in my_list_2[1]:
                            for j in my_list_2[1].split('/'):
                                roster.append([[couple_time],
                                               [my_list_2[3], f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}',
                                                dates1, couple_time, 9, j]])
                        else:
                            roster.append([[couple_time],
                                           [my_list_2[3], f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}', dates1,
                                            couple_time, 7, my_list_2[1]]])
                dates = str(datetime.now())
                dates = dates.split()
                dates = dates[-1].split(':')
                now_time = f'{int(dates[0])}:{dates[1]}'
            for i in roster:
                time = i[0][0].split('-')
                if now_time <= time[1]:
                    groups = i[1][-1].strip()
                    if groups in name_groups:
                        New_Events(f'{i[1][0]}', f'Время: {i[1][1]}', f'{i[1][2]}', f'{i[1][3]}', f'9', groups)
                    else:
                        start_message(TG_Bot_ID, f'Таких данных нет парсинг планшетки планшетки(2) \n'
                                                 f'Данные которых нет в списке: \'{groups}\'')
            for i in roster:
                time = i[0][0].split('-')
                if now_time <= time[1]:
                    teacher = i[1][1].split('. ')
                    teacher = teacher[0].split(': ')
                    teacher = f'{teacher[1]}.'

                    try:
                        text = i[1][1].split(': ')
                        text1 = text[-1].split('. ')
                        teacher = teacher.strip()
                        New_Events(f'{i[1][0]}', f'Время: {text[0]}: {i[1][-1]} {text1[-1]}', f'{i[1][2]}',
                                   f'{i[1][3]}', f'7', f'{teacher}')
                    except Exception as e:
                        start_message(TG_Bot_ID, f'Была вызвана ошибка в разделе парсинга планшетки(2) \n'
                                                 f'Данные вызванные ошибку: {i}'
                                                 f'описание ошибки: {e}')
            start_message(TG_Bot_ID, f'Расписание с планшетки было занесено в календарь')
        except Exception as e:
            start_message(TG_Bot_ID, f'Была вызвана ошибка в разделе парсинга планшетки(2.1) \n'
                                     f'описание ошибки: {e}')

    def planchette_check():
        """Функция проверяет соответствует ли планшетка требованиям"""
        wb1 = openpyxl.load_workbook('planchette.xlsx')
        worksheet = wb1['1 пара'].max_column
        if worksheet == 9:
            Parsing.planshet()
        else:
            start_message(TG_Bot_ID, f'Планшетка не соответствует требованиям')


def Monday():
    for time in ['04:30', '07:00', '08:00', '09:00', '10:30', '12:20', '13:10', '16:20', '17:45']:
        schedule.every().day.at(time).do(planchet_id)


def Tuesday():
    for time in ['04:30', '07:00', '08:00', '09:00', '10:30', '12:20', '14:00', '15:50', '17:20', '18:00']:
        schedule.every().day.at(time).do(planchet_id)


def main():
    """Функция запускает другие функции в определенное время и день недели"""
    schedule.every().monday.at('04:25').do(Monday)
    schedule.every().tuesday.at('04:25').do(Tuesday)
    schedule.every().wednesday.at('04:24').do(Tuesday)
    schedule.every().thursday.at('04:24').do(Tuesday)
    schedule.every().friday.at('04:25').do(Tuesday)
    schedule.every().saturday.at('04:25').do(Tuesday)
    schedule.every().sunday.at('12:55').do(Parsing.Par_Group)
    schedule.every().sunday.at('14:00').do(Parsing.Par_Techer)
    while True:
        schedule.run_pending()


if __name__ == '__main__':
    planchet_id()

bot.infinity_polling()